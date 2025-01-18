import re
import pandas as pd
from openpyxl import load_workbook

def grade_assignment(code, html_path, excel_path):
    """
    Grades Assignment 3 based on the provided code, HTML file, and Excel file.
    Returns a numerical grade (0-100).
    """
    total_score = 0
    debug_info = []

    ##########################################
    # 1. Library Imports (10 Points)
    ##########################################
    imports_score = 0
    # Check for required libraries
    if re.search(r"(?i)(import|from)\s+(gspread|pygsheets|google\-api\-python\-client)\b", code):
        imports_score += 4  # Points for gspread, pygsheets, or google-api-python-client
    if re.search(r"(?i)(import|from)\s+(pandas|numpy)\b", code):
        imports_score += 2  # Points for pandas or numpy
    if re.search(r"(?i)(import|from)\s+(folium|plotly|geopandas|matplotlib)\b", code):
        imports_score += 4  # Points for folium, plotly, geopandas, or matplotlib with mapping extensions
    debug_info.append(f"Library Imports: {imports_score} / 10")
    total_score += imports_score

    ##########################################
    # 2. Code Quality (10 Points)
    ##########################################
    quality_score = 0
    # a) Descriptive Variable Names (5 Points)
    if re.search(r"\b(temperature|temp|longitude|lat|map_data)\b", code, re.IGNORECASE):
        quality_score += 5
    # b) Spacing (5 Points)
    if not re.search(r"\S[=<>+\-/*]{1}\S", code):
        quality_score += 5
    debug_info.append(f"Code Quality: {quality_score} / 10")
    total_score += quality_score

    ##########################################
    # 3. Using JSON API (10 Points)
    ##########################################
    json_api_score = 0
    if re.search(r"(?i)(https?://\S+json\b)", code):
        json_api_score += 10
    debug_info.append(f"JSON API Usage: {json_api_score} / 10")
    total_score += json_api_score

    ##########################################
    # 4. Encapsulate Functionality (5 Points)
    ##########################################
    encapsulation_score = 0
    if re.search(r"def\s+stage_1\b", code) and re.search(r"def\s+stage_2\b", code) and re.search(r"def\s+stage_3\b", code):
        encapsulation_score += 5
    debug_info.append(f"Encapsulation: {encapsulation_score} / 5")
    total_score += encapsulation_score

    ##########################################
    # 5. Filter Data Below 25°C (5 Points)
    ##########################################
    filter_below_score = 0
    if re.search(r"temperature\s*<\s*25", code):
        filter_below_score += 5
    debug_info.append(f"Filter Below 25°C: {filter_below_score} / 5")
    total_score += filter_below_score

    ##########################################
    # 6. Filter Data Above 25°C (5 Points)
    ##########################################
    filter_above_score = 0
    if re.search(r"temperature\s*>\s*25", code):
        filter_above_score += 5
    debug_info.append(f"Filter Above 25°C: {filter_above_score} / 5")
    total_score += filter_above_score

    ##########################################
    # 7. HTML File Grading (15 Points)
    ##########################################
    html_score = 0
    try:
        with open(html_path, "r", encoding="utf-8") as f:
            html_content = f.read().lower()
            # Check for markers
            if "marker(" in html_content or "circle-marker(" in html_content:
                html_score += 5
            # Check for colors
            if "green" in html_content:
                html_score += 5
            if "red" in html_content:
                html_score += 5
    except Exception as e:
        debug_info.append(f"HTML File Error: {e}")
    debug_info.append(f"HTML File: {html_score} / 15")
    total_score += html_score

    ##########################################
    # 8. Excel File Grading (25 Points)
    ##########################################
    excel_score = 0
    try:
        wb = load_workbook(excel_path)
        # Check sheet names
        sheet_names = wb.sheetnames
        if "Sheet1" in sheet_names:
            excel_score += 5
        if "Below_25" in sheet_names:
            excel_score += 5
        if "Above_25" in sheet_names:
            excel_score += 5

        # Check column names in Below_25 and Above_25 sheets
        for sheet in ["Below_25", "Above_25"]:
            if sheet in sheet_names:
                df = pd.read_excel(excel_path, sheet_name=sheet)
                columns = [col.lower() for col in df.columns]
                if "longitude" in columns and "latitude" in columns and ("temperature" in columns or "temp" in columns):
                    excel_score += 5

        # Check row counts
        if "Below_25" in sheet_names:
            df = pd.read_excel(excel_path, sheet_name="Below_25")
            if abs(len(df) - 264) <= 3:
                excel_score += 5
        if "Above_25" in sheet_names:
            df = pd.read_excel(excel_path, sheet_name="Above_25")
            if abs(len(df) - 237) <= 3:
                excel_score += 5
    except Exception as e:
        debug_info.append(f"Excel File Error: {e}")
    debug_info.append(f"Excel File: {excel_score} / 25")
    total_score += excel_score

    ##########################################
    # Final Score Calculation
    ##########################################
    total_score = min(total_score, 100)  # Ensure the score does not exceed 100
    debug_info.append(f"Total Score: {total_score} / 100")
    print("\n".join(debug_info))  # Debug info printed to the console
    return total_score
